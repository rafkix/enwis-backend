import io
import uuid
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from app.modules.questions.models import QuestionType

TEMPLATE_HEADERS = [
    "Question Text*",
    "Question Type*",
    "Points*",
    "Order",
    "Correct Answer",
    "Explanation",
    "Option 1",
    "Option 1 Correct",
    "Option 2",
    "Option 2 Correct",
    "Option 3",
    "Option 3 Correct",
    "Option 4",
    "Option 4 Correct",
]

VALID_TYPES = {t.value for t in QuestionType}


def generate_template() -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions Template"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, header in enumerate(TEMPLATE_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    ws.add_data_validation(
        _make_dropdown("B2:B1048576", list(VALID_TYPES))
    )
    ws.add_data_validation(
        _make_dropdown("H2:H1048576", ["TRUE", "FALSE"])
    )
    ws.add_data_validation(
        _make_dropdown("J2:J1048576", ["TRUE", "FALSE"])
    )
    ws.add_data_validation(
        _make_dropdown("L2:L1048576", ["TRUE", "FALSE"])
    )
    ws.add_data_validation(
        _make_dropdown("N2:N1048576", ["TRUE", "FALSE"])
    )

    _set_col_widths(ws, [40, 18, 8, 8, 30, 40, 40, 12, 40, 12, 40, 12, 40, 12])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _make_dropdown(ref: str, values: list[str]):
    return DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=True)


def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w


def _cell(row: tuple, col: int) -> Any:
    """Safely get a cell from a row tuple, returning None if out of range."""
    return row[col] if col < len(row) else None


def parse_excel(file_bytes: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    questions: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []

    for idx, row in enumerate(rows, start=2):
        if all(cell is None for cell in row):
            continue

        q_text = str(_cell(row, 0) or "").strip()
        q_type = str(_cell(row, 1) or "").strip().lower()
        points_raw = _cell(row, 2)
        order_raw = _cell(row, 3)
        correct_answer = str(_cell(row, 4) or "").strip() or None
        explanation = str(_cell(row, 5) or "").strip() or None

        if not q_text:
            errors.append({"row": idx, "error": "Question text is required"})
            continue
        if q_type not in VALID_TYPES:
            valid_types = ", ".join(sorted(VALID_TYPES))
            errors.append({"row": idx, "error": f"Invalid type '{q_type}'. Valid: {valid_types}"})
            continue

        try:
            points = int(points_raw) if points_raw is not None else 1
        except (ValueError, TypeError):
            errors.append({"row": idx, "error": "Points must be an integer"})
            continue

        order = None
        if order_raw is not None:
            try:
                order = int(order_raw)
            except (ValueError, TypeError):
                errors.append({"row": idx, "error": "Order must be an integer"})
                continue

        options = []
        for opt_idx in range(4):
            opt_text_raw = _cell(row, 6 + opt_idx * 2)
            opt_correct_raw = _cell(row, 7 + opt_idx * 2)
            opt_text = str(opt_text_raw or "").strip()
            if opt_text:
                is_correct = str(opt_correct_raw or "").strip().upper() == "TRUE"
                options.append({
                    "id": uuid.uuid4(),
                    "text": opt_text,
                    "is_correct": is_correct,
                    "order": opt_idx + 1,
                })

        question = {
            "text": q_text,
            "question_type": q_type,
            "points": points,
            "correct_answer": correct_answer,
            "explanation": explanation,
            "options": options,
        }
        if order is not None:
            question["order"] = order
        questions.append(question)

    return questions, errors

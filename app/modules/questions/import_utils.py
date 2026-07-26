"""Excel / CSV parsing for bulk Question import.

Produces dicts shaped exactly like ``QuestionRepository.create`` expects
(title, question_type, difficulty, score, explanation, choices=[...]),
so results can be handed straight to ``QuestionService.bulk_create_questions``
— no separate insertion path, no duplicated business logic.
"""

import csv
import io
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

TEMPLATE_HEADERS = [
    "Title*",
    "Question Type*",
    "Score",
    "Explanation",
    "Option 1",
    "Option 1 Correct",
    "Option 2",
    "Option 2 Correct",
    "Option 3",
    "Option 3 Correct",
    "Option 4",
    "Option 4 Correct",
]

VALID_TYPES = {"single_choice", "short_answer", "image"}


def generate_import_template() -> io.BytesIO:
    """Build a downloadable .xlsx template with a Question Type dropdown."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions Template"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col, header in enumerate(TEMPLATE_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    ws.add_data_validation(_dropdown("B2:B1048576", sorted(VALID_TYPES)))
    for col_letter in ("F", "H", "J", "L"):
        ws.add_data_validation(_dropdown(f"{col_letter}2:{col_letter}1048576", ["TRUE", "FALSE"]))

    for i, width in enumerate([35, 18, 8, 35, 30, 12, 30, 12, 30, 12, 30, 12], 1):
        ws.column_dimensions[chr(64 + i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _dropdown(ref: str, values: list[str]) -> DataValidation:
    dv = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=True)
    dv.add(ref)
    return dv


def _row_to_question(row: dict[str, Any], row_num: int, errors: list[dict[str, Any]]) -> dict | None:
    """Shared row->question mapping used by both the Excel and CSV parsers.

    ``row`` maps header name -> cell value (already stripped of whitespace).
    """
    title = str(row.get("Title*") or row.get("Title") or "").strip()
    qtype = str(row.get("Question Type*") or row.get("Question Type") or "").strip().lower()
    score_raw = row.get("Score")
    explanation = str(row.get("Explanation") or "").strip() or None

    if not title:
        errors.append({"row": row_num, "error": "Title is required"})
        return None
    if qtype not in VALID_TYPES:
        errors.append(
            {"row": row_num, "error": f"Invalid type '{qtype}'. Valid: {', '.join(sorted(VALID_TYPES))}"}
        )
        return None

    try:
        score = int(score_raw) if score_raw not in (None, "") else 1
    except (ValueError, TypeError):
        errors.append({"row": row_num, "error": "Score must be an integer"})
        return None

    choices = []
    for opt_idx in range(1, 5):
        text = str(row.get(f"Option {opt_idx}") or "").strip()
        if not text:
            continue
        is_correct = str(row.get(f"Option {opt_idx} Correct") or "").strip().upper() == "TRUE"
        choices.append({"content": text, "is_correct": is_correct, "order": opt_idx})

    return {
        "title": title,
        "question_type": qtype,
        "difficulty": "medium",
        "score": score,
        "explanation": explanation,
        "visibility": "private",
        "choices": choices,
    }


def parse_excel(file_bytes: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Parse an uploaded .xlsx file into (questions, errors)."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows_iter, [])]

    questions: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []

    for idx, raw_row in enumerate(rows_iter, start=2):
        if all(cell is None for cell in raw_row):
            continue
        row = dict(zip(header, raw_row, strict=False))
        question = _row_to_question(row, idx, errors)
        if question:
            questions.append(question)

    wb.close()
    return questions, errors


def parse_csv(file_bytes: bytes) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Parse an uploaded .csv file into (questions, errors). Expects the
    same column headers as the Excel template.
    """
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))

    questions: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []

    for idx, row in enumerate(reader, start=2):
        if not any((v or "").strip() for v in row.values()):
            continue
        question = _row_to_question(row, idx, errors)
        if question:
            questions.append(question)

    return questions, errors
